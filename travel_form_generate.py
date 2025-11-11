import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import io
from datetime import datetime, timedelta
from PIL import Image as PILImage, ImageDraw, ImageFont
import base64
import urllib.request
import re

# Page configuration
st.set_page_config(page_title="Georgetown Travel Form Generator", page_icon="✈️", layout="wide")

# Load Excel template
@st.cache_data
def load_excel_template():
    """Load the Excel template and identify form structure"""
    wb = openpyxl.load_workbook('Georgetown Domestic Travel Authorization Form.xlsx')
    ws = wb['Reimbursement Form']
    return wb, ws

def chunk_list(items, chunk_size):
    """Yield successive chunks of size chunk_size from items."""
    if items is None:
        return []
    return [items[i:i+chunk_size] for i in range(0, len(items), chunk_size)]

def pad_to_length(items, length, pad_value=''):
    """Return a copy padded to given length."""
    items = list(items)
    if len(items) < length:
        items.extend([pad_value] * (length - len(items)))
    return items

def number_text_input(label, key, value=0.0, min_value=0.0, placeholder="0.00"):
    """Text input that accepts numeric values only, with validation.
    Returns the numeric value and shows inline warnings if invalid."""
    # Initialize session state if not exists
    if key not in st.session_state:
        st.session_state[key] = str(value) if value else ""
    
    # Track validation state for this specific input
    validation_key = f"{key}_has_error"
    
    text_val = st.text_input(label, key=key, placeholder=placeholder)
    
    # If empty, return 0.0 (no validation needed, clear any previous errors)
    if not text_val or not text_val.strip():
        st.session_state[validation_key] = False
        return 0.0
    
    # Try to extract numeric value from input
    # Remove common non-numeric characters like $, commas, spaces, etc.
    cleaned_text = text_val.strip().replace('$', '').replace(',', '').replace(' ', '')
    
    # Try to parse as float
    has_error = False
    error_message = None
    try:
        num_val = float(cleaned_text)
        if num_val < min_value:
            num_val = min_value
        # Input is valid, clear error state
        st.session_state[validation_key] = False
        return num_val
    except (ValueError, AttributeError):
        # Check if there are any invalid characters
        # Allow: integers, decimals, negative numbers
        if not re.match(r'^-?\d+(\.\d+)?$', cleaned_text):
            has_error = True
            error_message = "⚠️ Invalid input. Please enter a valid number."
    
    # Update error state
    st.session_state[validation_key] = has_error
    
    # Show warning inline if there's an error
    if has_error and error_message:
        st.warning(error_message)
        return 0.0
    
    return 0.0

def generate_signature_image(text, width=600, height=120, scale_factor=3):
    """Generate a signature-style image from text with high resolution"""
    if not text or not text.strip():
        return None
    
    # Use scale factor for high-resolution rendering (render at 3x, then scale down)
    scaled_width = width * scale_factor
    scaled_height = height * scale_factor
    
    # Create an image with white background (blank/transparent-looking) at high resolution
    img = PILImage.new('RGB', (scaled_width, scaled_height), (255, 255, 255))
    draw = ImageDraw.Draw(img)
    
    # Try to use a cursive/signature-style font, fallback to default
    # Include more common paths and try PIL's built-in fonts
    font_paths = [
        '/System/Library/Fonts/Supplemental/SnellRoundhand.ttc',  # macOS
        '/System/Library/Fonts/Supplemental/Chalkduster.ttf',    # macOS alternative
        'C:/Windows/Fonts/brushsc.ttf',                           # Windows
        'C:/Windows/Fonts/BRUSHSCI.TTF',                          # Windows
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Oblique.ttf', # Linux
        '/usr/share/fonts/truetype/liberation/LiberationSans-Italic.ttf', # Linux alternative
        '/usr/share/fonts/truetype/noto/NotoSans-Italic.ttf',     # Linux alternative
    ]
    
    # Start with larger font size (scale it too)
    font_size = 72 * scale_factor  # Larger base font size
    font = None
    font_path_used = None
    
    for font_path in font_paths:
        try:
            font = ImageFont.truetype(font_path, font_size)
            font_path_used = font_path
            break
        except (OSError, IOError, Exception):
            continue
    
    # If no system font found, try to use PIL's default font or create a simple signature style
    if font is None:
        try:
            # Try common font names that might be available
            import platform
            system = platform.system()
            if system == 'Windows':
                # Try Windows common fonts
                for font_name in ['arial', 'calibri', 'times']:
                    try:
                        font = ImageFont.truetype(f"{font_name}.ttf", font_size)
                        font_path_used = font_name
                        break
                    except:
                        continue
            elif system == 'Linux':
                # Try Linux common fonts
                for font_path_linux in [
                    '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
                    '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                ]:
                    try:
                        font = ImageFont.truetype(font_path_linux, font_size)
                        font_path_used = font_path_linux
                        break
                    except:
                        continue
        except:
            pass
        
        # Final fallback: use PIL's default font
        if font is None:
            try:
                # Try to load default font with larger size
                font = ImageFont.load_default()
                font_size = 36 * scale_factor
            except:
                # Ultimate fallback
                font = ImageFont.load_default()
                font_size = 36 * scale_factor
    
    # Calculate text dimensions first
    try:
        bbox = draw.textbbox((0, 0), text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
    except:
        # Fallback if textbbox fails
        text_width = len(text) * font_size * 0.6
        text_height = font_size * 1.2
    
    # Adjust font size if text is too wide to fit in available width
    min_font_size = 30 * scale_factor
    while text_width > scaled_width - (40 * scale_factor) and font_size > min_font_size:
        font_size -= 3 * scale_factor
        try:
            if font_path_used and font_path_used not in ['arial', 'calibri', 'times']:
                font = ImageFont.truetype(font_path_used, font_size)
            else:
                font = ImageFont.load_default()
        except:
            font = ImageFont.load_default()
        try:
            bbox = draw.textbbox((0, 0), text, font=font)
            text_width = bbox[2] - bbox[0]
            text_height = bbox[3] - bbox[1]
        except:
            text_width = len(text) * font_size * 0.6
            text_height = font_size * 1.2
    
    # Calculate position (left-aligned with padding, vertically centered)
    padding = 20 * scale_factor
    x = padding
    y = (scaled_height - text_height) / 2
    
    # Draw the signature text in black with antialiasing
    try:
        draw.text((x, y), text, fill=(0, 0, 0), font=font)
    except Exception as e:
        # If font drawing fails, try with default font
        try:
            font = ImageFont.load_default()
            draw.text((x, y), text, fill=(0, 0, 0), font=font)
            # Recalculate dimensions with default font
            try:
                bbox = draw.textbbox((0, 0), text, font=font)
                text_width = bbox[2] - bbox[0]
                text_height = bbox[3] - bbox[1]
            except:
                text_width = len(text) * font_size * 0.6
                text_height = font_size * 1.2
        except:
            # Ultimate fallback - draw text without font specification
            draw.text((x, y), text, fill=(0, 0, 0))
    
    # Add a thicker underline for signature effect (also scaled)
    line_y = y + text_height + (8 * scale_factor)
    line_width = 3 * scale_factor
    draw.line([(x - 5 * scale_factor, line_y), (x + text_width + 5 * scale_factor, line_y)], 
              fill=(0, 0, 0), width=int(line_width))
    
    # Calculate the actual bounds including underline
    actual_bottom = line_y + 5 * scale_factor
    actual_right = x + text_width + 30 * scale_factor
    
    # Crop to actual content with some padding, but ensure full signature is visible
    img = img.crop((0, 0, min(scaled_width, max(int(actual_right), int(text_width) + padding * 2)), 
                    min(scaled_height, max(int(actual_bottom), int(text_height) + padding * 2))))
    
    # Scale down using high-quality resampling for sharp, clear output
    final_width = img.size[0] // scale_factor
    final_height = img.size[1] // scale_factor
    img = img.resize((final_width, final_height), PILImage.Resampling.LANCZOS)
    
    return img

def generate_date_range(start_date, end_date, max_days=7):
    """Generate a list of dates from start_date to end_date, formatted as MM/DD/YY"""
    if not start_date or not end_date:
        return [''] * max_days
    
    dates = []
    current_date = start_date
    while current_date <= end_date and len(dates) < max_days:
        # Format as MM/DD/YY
        formatted_date = current_date.strftime('%m/%d/%y')
        dates.append(formatted_date)
        current_date += timedelta(days=1)
    
    # Fill remaining slots with empty strings
    while len(dates) < max_days:
        dates.append('')
    
    return dates

def get_red_cells(ws):
    """Identify all cells that should be filled (red highlighted)"""
    red_cells = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                fill_color = cell.fill.start_color.rgb
                if fill_color == 'FFFF0000':  # Red color
                    cell_ref = cell.coordinate
                    row_num = cell.row
                    col_num = cell.column
                    if cell.value is None or str(cell.value).strip() == "" or str(cell.value) == "None":
                        red_cells[cell_ref] = {
                            'row': row_num,
                            'col': col_num,
                            'value': None
                        }
                    else:
                        # Some red cells have formulas or values, we'll track them too
                        red_cells[cell_ref] = {
                            'row': row_num,
                            'col': col_num,
                            'value': cell.value
                        }
    return red_cells

def create_pdf(form_data, ws):
    """Create PDF with form data and red highlighting"""
    meal_deductions = {
        68: { 'breakfast': 16, 'lunch': 19, 'dinner': 28, 'incidental': 5, 'first_last': 51.00 },
        74: { 'breakfast': 18, 'lunch': 20, 'dinner': 31, 'incidental': 5, 'first_last': 55.50 },
        80: { 'breakfast': 20, 'lunch': 22, 'dinner': 33, 'incidental': 5, 'first_last': 60.00 },
        86: { 'breakfast': 22, 'lunch': 23, 'dinner': 36, 'incidental': 5, 'first_last': 64.50 },
        92: { 'breakfast': 23, 'lunch': 26, 'dinner': 38, 'incidental': 5, 'first_last': 69.00 },
    }
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, 
                            rightMargin=0.5*inch, leftMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#000000'),
        spaceAfter=12,
        alignment=1  # Center
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#000000'),
        spaceAfter=12,
        alignment=1  # Center
    )
    
    # Helper to load, trim and whiten logo from URL and size by target height
    def load_logo_image(url: str, target_height_inch: float):
        try:
            with urllib.request.urlopen(url) as resp:
                data = resp.read()
            img_pil = PILImage.open(io.BytesIO(data)).convert('RGB')
            # Replace near-black background with white and trim borders
            pixels = img_pil.load()
            width, height = img_pil.size
            # Replace very dark pixels with white to avoid giant black boxes
            for y in range(height):
                for x in range(width):
                    r, g, b = pixels[x, y]
                    if r < 20 and g < 20 and b < 20:
                        pixels[x, y] = (255, 255, 255)
            # Create trim mask for white background to crop extra whitespace
            gray = img_pil.convert('L')
            # Inverse mask of non-white areas
            mask = gray.point(lambda p: 0 if p > 250 else 255)
            bbox = mask.getbbox()
            if bbox:
                img_pil = img_pil.crop(bbox)
            # Scale by target height
            target_h = target_height_inch * inch
            w, h = img_pil.size
            aspect = w / h if h else 1.0
            target_w = target_h * aspect
            buf = io.BytesIO()
            img_pil.save(buf, format='PNG')
            buf.seek(0)
            return Image(buf, width=target_w, height=target_h)
        except Exception:
            return None

    # Logos beside title
    georgetown_logo_url = 'https://raw.githubusercontent.com/JiaqinWu/HRSA64_Dash/main/Georgetown_logo_blueRGB.png'
    advance_logo_url = 'https://raw.githubusercontent.com/JiaqinWu/HRSA64_Dash/main/ADVANCE%20Logo_Horizontal%20Blue.png'

    left_logo = load_logo_image(georgetown_logo_url, target_height_inch=0.8)
    right_logo = load_logo_image(advance_logo_url, target_height_inch=0.4)

    title_para = Paragraph("Domestic Travel Authorization Form", title_style)

    title_block = [[title_para]]
    title_table = Table(title_block)
    title_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    # Build a clean 3-column header row
    # Reserve ~1.3in for each logo, center column takes remaining width
    content_width = 8.5*inch - (0.5*inch + 0.5*inch)
    left_w = 1.3*inch
    right_w = 1.0*inch
    center_w = max(content_width - (left_w + right_w), 3.5*inch)
    header_row = [left_logo if left_logo else '', title_table, right_logo if right_logo else '']
    header_table = Table([header_row], colWidths=[left_w, center_w, right_w])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    story.append(header_table)
    # Thin rule below header
    story.append(Spacer(1, 0.05*inch))
    story.append(Table([["" ]], colWidths=[content_width], rowHeights=[0.5]))    
    story.append(Spacer(1, 0.1*inch))
    
    # Traveler Information Section
    story.append(Paragraph("<b>Traveler Information</b>", styles['Heading2']))
    story.append(Spacer(1, 0.1*inch))
    
    # Create traveler info table
    traveler_data = [
        ['Name', form_data.get('name', ''), 'Organization', form_data.get('organization', 'Georgetown University')],
        ['Address Line 1', form_data.get('address1', ''), 'Destination', form_data.get('destination', '')],
        ['Address Line 2', form_data.get('address2', ''), 'Departure Date', form_data.get('departure_date', '')],
        ['City', form_data.get('city', ''), 'Return Date', form_data.get('return_date', '')],
        ['State', form_data.get('state', ''), 'Email Address', form_data.get('email', '')],
        ['Zip', form_data.get('zip', ''), '', '']
    ]
    
    traveler_table = Table(traveler_data, colWidths=[1.5*inch, 1.8*inch, 1.5*inch, 1.8*inch])
    traveler_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E0E0E0')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#E0E0E0')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.red),  # Red text for input fields
        ('TEXTCOLOR', (3, 0), (3, -1), colors.red),  # Red text for input fields
        ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#FFEBEE')),  # Light red background
        ('BACKGROUND', (3, 0), (3, -1), colors.HexColor('#FFEBEE')),  # Light red background
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(traveler_table)
    story.append(Spacer(1, 0.15*inch))
    # Traveler Paid Expenses Section
    story.append(Paragraph("<b>Purpose of Travel</b>", styles['Heading2']))
    story.append(Spacer(1, 0.15*inch))
    
    # Traveler Paid Expenses Section
    story.append(Paragraph("<b>Traveler Paid Expenses</b>", styles['Heading2']))
    
    # Mileage Section
    story.append(Paragraph("<b>Mileage</b>", styles['Heading3']))
    story.append(Paragraph("Mileage for 2025 is $0.70 per mile.", styles['Normal']))
    story.append(Spacer(1, 0.1*inch))
    
    # Mileage: build multiple tables, 7 days per table
    all_mileage_dates = form_data.get('mileage_dates', [])
    all_mileage_amounts = form_data.get('mileage_amounts', [])
    # Grand total for mileage rate across all days
    grand_mileage_rate_total = 0.0
    for amount in all_mileage_amounts:
        if amount and str(amount).strip():
            try:
                grand_mileage_rate_total += round(float(amount) * 0.70, 0)
            except:
                pass
    grand_mileage_rate_total = round(grand_mileage_rate_total, 0)
    mileage_tables = []
    mileage_dates_chunks = chunk_list(all_mileage_dates, 7)
    mileage_amount_chunks = chunk_list(all_mileage_amounts, 7)
    total_mileage_chunks = len(mileage_dates_chunks)
    for idx in range(total_mileage_chunks):
        dates_chunk = mileage_dates_chunks[idx] if idx < len(mileage_dates_chunks) else []
        amounts_chunk = mileage_amount_chunks[idx] if idx < len(mileage_amount_chunks) else []
        dates_chunk = pad_to_length(dates_chunk, 7, '')
        amounts_chunk = pad_to_length(amounts_chunk, 7, '')
        mileage_data = [['Date (MM/DD/YY)'] + dates_chunk + ['Total']]
        mileage_data.append(['MILEAGE (Per Day)'] + [str(x) if x else '' for x in amounts_chunk] + [''])
        # Calculate mileage rates per chunk
        mileage_rates = []
        for amount in amounts_chunk:
            if amount and str(amount).strip():
                try:
                    rate = round(float(amount) * 0.70, 0)
                    mileage_rates.append(f"${int(rate)}")
                except:
                    mileage_rates.append('')
            else:
                mileage_rates.append('')
        # Only last table shows grand total; others blank
        mileage_total_cell = f"${int(grand_mileage_rate_total)}" if idx == total_mileage_chunks - 1 else ''
        mileage_data.append(['Mileage Rate'] + mileage_rates + [mileage_total_cell])
        mileage_table = Table(mileage_data, colWidths=[1.3*inch] + [0.7*inch]*7 + [0.75*inch])
        mileage_table.setStyle(TableStyle([
            # Headers and date row set to white for print
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('BACKGROUND', (1, 0), (7, 0), colors.white),
            ('TEXTCOLOR', (1, 0), (7, 0), colors.black),
            ('BACKGROUND', (0, 1), (0, 1), colors.HexColor('#E0E0E0')),
            ('TEXTCOLOR', (1, 1), (7, 1), colors.red), 
            ('BACKGROUND', (1, 1), (7, 1), colors.HexColor('#FFF5F5')),
            ('BACKGROUND', (0, 2), (0, 2), colors.HexColor('#E0E0E0')),
            ('TEXTCOLOR', (1, 2), (7, 2), colors.red), 
            ('BACKGROUND', (1, 2), (7, 2), colors.HexColor('#FFF5F5')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('TEXTCOLOR', (8, 2), (8, 2), colors.red),
            ('BACKGROUND', (8, 2), (8, 2), colors.HexColor('#FFF5F5')),
        ]))
        mileage_tables.append(mileage_table)
    for t in mileage_tables:
        story.append(t)
        story.append(Spacer(1, 0.15*inch))
    
    # Expenses Section - 7 days + total column
    story.append(Paragraph("<b>Airfare, Transportation, Parking, Lodging, Miscellaneous.</b>", styles['Heading3']))
    story.append(Paragraph("Ground Transportation Includes: Taxi, Uber, etc.", styles['Normal']))
    story.append(Paragraph("Miscellaneous/Other: Pre-approved travel expenses not listed in this form", styles['Normal']))
    story.append(Spacer(1, 0.1*inch))
    expense_dates = form_data.get('expense_dates', [])
    airfare = form_data.get('airfare', [])
    ground_transport = form_data.get('ground_transport', [])
    parking = form_data.get('parking', [])
    lodging = form_data.get('lodging', [])
    baggage = form_data.get('baggage', [])
    misc = form_data.get('misc', [])
    misc2 = form_data.get('misc2', [])  # Second row for misc expenses
    
    # Build labels; only use if descriptions are actually provided
    misc_desc1_val = form_data.get('misc_desc1', '').strip() if form_data.get('misc_desc1', '') else ''
    misc_desc2_val = form_data.get('misc_desc2', '').strip() if form_data.get('misc_desc2', '') else ''
    
    # Only show misc rows that have actual descriptions
    misc_label1 = misc_desc1_val if misc_desc1_val else None
    misc_label2 = misc_desc2_val if misc_desc2_val else None
    
    # Grand totals across all days
    grand_af = sum(x for x in airfare if x)
    grand_gt = sum(x for x in ground_transport if x)
    grand_pk = sum(x for x in parking if x)
    grand_lg = sum(x for x in lodging if x)
    grand_bg = sum(x for x in baggage if x)
    grand_m1 = sum(x for x in misc if x)
    grand_m2 = sum(x for x in misc2 if x)

    expense_tables = []
    expense_chunks = chunk_list(expense_dates, 7)
    total_expense_chunks = len(expense_chunks)
    for i, dates_chunk in enumerate(expense_chunks):
        chunk_len = len(pad_to_length(dates_chunk, 7))
        pad_len = 7
        af = pad_to_length(airfare[i*7:(i+1)*7], pad_len, 0)
        gt = pad_to_length(ground_transport[i*7:(i+1)*7], pad_len, 0)
        pk = pad_to_length(parking[i*7:(i+1)*7], pad_len, 0)
        lg = pad_to_length(lodging[i*7:(i+1)*7], pad_len, 0)
        bg = pad_to_length(baggage[i*7:(i+1)*7], pad_len, 0)
        m1 = pad_to_length(misc[i*7:(i+1)*7], pad_len, 0)
        m2 = pad_to_length(misc2[i*7:(i+1)*7], pad_len, 0)
        # Build expenses data - only include misc rows if descriptions are provided
        expenses_data = [
            ['Date (MM/DD/YY)'] + pad_to_length(dates_chunk, 7, '') + ['Total'],
            ['Airfare'] + [f"${x:.2f}" if x else '' for x in af] + ([f"${grand_af:.2f}"] if i == total_expense_chunks - 1 else ['']),
            ['Ground Transportation'] + [f"${x:.2f}" if x else '' for x in gt] + ([f"${grand_gt:.2f}"] if i == total_expense_chunks - 1 else ['']),
            ['Parking'] + [f"${x:.2f}" if x else '' for x in pk] + ([f"${grand_pk:.2f}"] if i == total_expense_chunks - 1 else ['']),
            ['Lodging'] + [f"${x:.2f}" if x else '' for x in lg] + ([f"${grand_lg:.2f}"] if i == total_expense_chunks - 1 else ['']),
            ['Baggage Fees'] + [f"${x:.2f}" if x else '' for x in bg] + ([f"${grand_bg:.2f}"] if i == total_expense_chunks - 1 else ['']),
            ['Miscellaneous/Other\n(Provide Description)'] + [''] * 7 + [''],
        ]
        
        # Only add misc rows if descriptions are provided
        misc_row1_idx = None
        misc_row2_idx = None
        if misc_label1 is not None:
            expenses_data.append([misc_label1] + [f"${x:.2f}" if x else '' for x in m1] + ([f"${grand_m1:.2f}"] if i == total_expense_chunks - 1 else ['']))
            misc_row1_idx = len(expenses_data) - 1
        if misc_label2 is not None:
            expenses_data.append([misc_label2] + [f"${x:.2f}" if x else '' for x in m2] + ([f"${grand_m2:.2f}"] if i == total_expense_chunks - 1 else ['']))
            misc_row2_idx = len(expenses_data) - 1
        
        expenses_table = Table(expenses_data, colWidths=[1.3*inch] + [0.65*inch]*7 + [0.75*inch])
        
        # Build table style - dynamically handle misc rows
        table_style = [
            # Left label column light gray; headers white
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E0E0E0')),
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('BACKGROUND', (1, 0), (7, 0), colors.white),
            ('TEXTCOLOR', (1, 0), (7, 0), colors.black),
            ('TEXTCOLOR', (1, 1), (7, 5), colors.red),
            ('BACKGROUND', (1, 1), (7, 5), colors.HexColor('#FFF5F5')),
            ('SPAN', (0, 6), (7, 6)),  # Span across all day columns
            ('TEXTCOLOR', (0, 6), (0, 6), colors.black),
            ('BACKGROUND', (0, 6), (7, 6), colors.white),
        ]
        
        # Add styling for misc rows only if they exist
        if misc_row1_idx is not None:
            table_style.extend([
                ('TEXTCOLOR', (0, misc_row1_idx), (0, misc_row1_idx), colors.black),
                ('BACKGROUND', (0, misc_row1_idx), (0, misc_row1_idx), colors.HexColor('#E0E0E0')),
                ('TEXTCOLOR', (1, misc_row1_idx), (7, misc_row1_idx), colors.red),
                ('BACKGROUND', (1, misc_row1_idx), (7, misc_row1_idx), colors.HexColor('#FFF5F5')),
                ('TEXTCOLOR', (8, misc_row1_idx), (8, misc_row1_idx), colors.red),
                ('BACKGROUND', (8, misc_row1_idx), (8, misc_row1_idx), colors.HexColor('#FFF5F5')),
            ])
        if misc_row2_idx is not None:
            table_style.extend([
                ('TEXTCOLOR', (0, misc_row2_idx), (0, misc_row2_idx), colors.black),
                ('BACKGROUND', (0, misc_row2_idx), (0, misc_row2_idx), colors.HexColor('#E0E0E0')),
                ('TEXTCOLOR', (1, misc_row2_idx), (7, misc_row2_idx), colors.red),
                ('BACKGROUND', (1, misc_row2_idx), (7, misc_row2_idx), colors.HexColor('#FFF5F5')),
                ('TEXTCOLOR', (8, misc_row2_idx), (8, misc_row2_idx), colors.red),
                ('BACKGROUND', (8, misc_row2_idx), (8, misc_row2_idx), colors.HexColor('#FFF5F5')),
            ])
        
        # Add common styling for totals column
        last_row = len(expenses_data) - 1
        table_style.extend([
            ('TEXTCOLOR', (8, 1), (8, last_row), colors.red),
            ('BACKGROUND', (8, 1), (8, last_row), colors.HexColor('#FFF5F5')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (7, -1), 'CENTER'),
            ('ALIGN', (8, 0), (8, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
        ])
        
        expenses_table.setStyle(TableStyle(table_style))
        expense_tables.append(expenses_table)
    for t in expense_tables:
        story.append(t)
        story.append(Spacer(1, 0.15*inch))
    
    # Meals and Incidentals Section
    story.append(Paragraph("<b>Meals and Incidentals Per Diem</b>", styles['Heading2']))
    story.append(Paragraph("Federal Guidelines: On the first and last travel day, travelers are only eligible for 75 percent of the total M&IE rate.", styles['Normal']))
    story.append(Spacer(1, 0.1*inch))
    
    per_diem_dates = form_data.get('per_diem_dates', [])
    per_diem_amounts = form_data.get('per_diem_amounts', [])  # Will be one of PER_DIEM_OPTIONS
    breakfast_checks = form_data.get('breakfast_checks', [])
    lunch_checks = form_data.get('lunch_checks', [])
    dinner_checks = form_data.get('dinner_checks', [])
    
    # Calculate adjusted per diem for each day using dollar-based deductions
    adjusted_per_diem = []
    daily_totals = []
    
    # Find which days have dates (non-empty)
    days_with_dates = [i for i, d in enumerate(per_diem_dates) if d and str(d).strip()]
    num_days = len(days_with_dates)
    first_day_idx = days_with_dates[0] if days_with_dates else 0
    last_day_idx = days_with_dates[-1] if days_with_dates else 0
    
    for i in range(len(per_diem_dates)):
        if i < len(per_diem_dates) and per_diem_dates[i] and str(per_diem_dates[i]).strip():
            base_per_diem = int(per_diem_amounts[i]) if (i < len(per_diem_amounts) and per_diem_amounts[i]) else 80
            deducts = meal_deductions.get(base_per_diem, meal_deductions[80])
            deduction_total = 0.0
            if i < len(breakfast_checks) and breakfast_checks[i]:
                deduction_total += deducts['breakfast']
            if i < len(lunch_checks) and lunch_checks[i]:
                deduction_total += deducts['lunch']
            if i < len(dinner_checks) and dinner_checks[i]:
                deduction_total += deducts['dinner']
            # Base already includes incidentals; do not add +$5 here
            pre75_total = max(0.0, float(base_per_diem) - deduction_total)
            # Apply 75% for first and last day
            if i == first_day_idx or i == last_day_idx:
                final_per_diem = round(pre75_total * 0.75, 2)
            else:
                final_per_diem = round(pre75_total, 2)
            
            adjusted_per_diem.append(final_per_diem)
            daily_totals.append(final_per_diem)
        else:
            adjusted_per_diem.append(0.0)
            daily_totals.append(0.0)
    
    total_per_diem_calculated = sum(daily_totals)
    
    # Calculate daily meal totals (before 75% reduction) and total dollar reductions
    daily_meal_totals = []
    total_reductions = []
    for i in range(len(per_diem_dates)):
        if i < len(per_diem_dates) and per_diem_dates[i] and str(per_diem_dates[i]).strip():
            base_per_diem = int(per_diem_amounts[i]) if per_diem_amounts[i] else 80
            deducts = meal_deductions.get(base_per_diem, meal_deductions[80])
            deduction_total = 0.0
            if i < len(breakfast_checks) and breakfast_checks[i]:
                deduction_total += deducts['breakfast']
            if i < len(lunch_checks) and lunch_checks[i]:
                deduction_total += deducts['lunch']
            if i < len(dinner_checks) and dinner_checks[i]:
                deduction_total += deducts['dinner']
            total_reductions.append(round(deduction_total, 2))
            pre75_total = max(0.0, float(base_per_diem) - deduction_total)
            daily_meal_totals.append(round(pre75_total, 2))
        else:
            total_reductions.append(0.0)
            daily_meal_totals.append(0.0)
    
    # Build per diem tables per 7-day chunk
    per_diem_tables = []
    for i, dates_chunk in enumerate(chunk_list(per_diem_dates, 7)):
        idx_start = i * 7
        pd = pad_to_length(dates_chunk, 7, '')
        amounts = pad_to_length(per_diem_amounts[idx_start:idx_start+7], 7, 80)
        bchk = pad_to_length(breakfast_checks[idx_start:idx_start+7], 7, False)
        lchk = pad_to_length(lunch_checks[idx_start:idx_start+7], 7, False)
        dchk = pad_to_length(dinner_checks[idx_start:idx_start+7], 7, False)
        # Map totals slice
        red = pad_to_length(total_reductions[idx_start:idx_start+7], 7, 0.0)
        meal_tot = pad_to_length(daily_meal_totals[idx_start:idx_start+7], 7, 0.0)
        adj = pad_to_length(adjusted_per_diem[idx_start:idx_start+7], 7, 0.0)
        # Determine labels for deductions in this chunk: use the per diem amount from actual dates
        # Find the first per diem amount that corresponds to a date (non-empty date)
        common_amount = None
        for j in range(len(pd)):
            if pd[j] and str(pd[j]).strip() and j < len(amounts) and amounts[j]:
                try:
                    common_amount = int(amounts[j])
                    break
                except (ValueError, TypeError):
                    continue
        # If no date found in this chunk, try to get from the actual per_diem_amounts (not padded)
        if common_amount is None and idx_start < len(per_diem_amounts):
            for j in range(idx_start, min(idx_start + 7, len(per_diem_amounts))):
                if j < len(per_diem_dates) and per_diem_dates[j] and str(per_diem_dates[j]).strip():
                    try:
                        common_amount = int(per_diem_amounts[j])
                        break
                    except (ValueError, TypeError):
                        continue
        # Default to 80 if still not found
        if common_amount is None:
            common_amount = 80
        
        if common_amount in meal_deductions:
            b_lbl = f"Breakfast -${meal_deductions[common_amount]['breakfast']}"
            l_lbl = f"Lunch -${meal_deductions[common_amount]['lunch']}"
            d_lbl = f"Dinner -${meal_deductions[common_amount]['dinner']}"
        else:
            b_lbl = "Breakfast -$"
            l_lbl = "Lunch -$"
            d_lbl = "Dinner -$"

        per_diem_data = [
            ['Date (MM/DD/YY)'] + [d if d and str(d).strip() else '' for d in pd] + [''],
            ['Per Diem Allowance'] + [f"${int(x)}" if (x and pd[j]) else '' for j, x in enumerate(amounts)] + [''],
            ['ADJUSTED PER DIEM', 'If meals were provided by Georgetown University (Place "x" in box)', '', '', '', '', '', ''],
            [b_lbl] + ['X' if (bchk[j] and pd[j]) else '' for j in range(7)] + [''],
            [l_lbl] + ['X' if (lchk[j] and pd[j]) else '' for j in range(7)] + [''],
            [d_lbl] + ['X' if (dchk[j] and pd[j]) else '' for j in range(7)] + [''],
            ['Total Reduction ($)'] + [f"${x:.2f}" if x != 0 else '' for x in red] + [''],
            ['Daily Meal Total'] + [f"${x:.2f}" if x > 0 else '' for x in meal_tot] + [''],
            ['Total Per Diem'] + [f"${x:.2f}" if x > 0 else '' for x in adj] + [f"${total_per_diem_calculated:.2f}" if i == len(chunk_list(per_diem_dates,7)) - 1 else ''],
        ]
        per_diem_table = Table(per_diem_data, colWidths=[1.3*inch] + [0.6*inch]*7 + [0.75*inch])
        per_diem_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E0E0E0')),
            # Date row and header white
            ('BACKGROUND', (1, 0), (7, 0), colors.white),
            ('TEXTCOLOR', (1, 0), (7, 0), colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('TEXTCOLOR', (1, 1), (-1, -1), colors.red),
            ('BACKGROUND', (1, 1), (-1, -1), colors.HexColor('#FFF5F5')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('SPAN', (0, 2), (7, 2)),
            ('BACKGROUND', (0, 2), (7, 2), colors.HexColor('#FFF5F5')),
            ('TEXTCOLOR', (0, 2), (7, 2), colors.red),
            ('TEXTCOLOR', (8, -1), (8, -1), colors.red),
            ('BACKGROUND', (8, -1), (8, -1), colors.HexColor('#FFF5F5')),
        ]))
        per_diem_tables.append(per_diem_table)
    for t in per_diem_tables:
        story.append(t)
        story.append(Spacer(1, 0.15*inch))
    
    # Totals Section
    total_mileage = form_data.get('total_mileage', 0)
    total_airfare = form_data.get('total_airfare', 0)
    total_ground_transport = form_data.get('total_ground_transport', 0)
    total_parking = form_data.get('total_parking', 0)
    total_lodging = form_data.get('total_lodging', 0)
    total_baggage = form_data.get('total_baggage', 0)
    total_misc = form_data.get('total_misc', 0)
    total_per_diem = form_data.get('total_per_diem', 0)
    
    # Calculate subtotal
    subtotal = total_mileage + total_airfare + total_ground_transport + total_parking + total_lodging + total_baggage + total_misc + total_per_diem
    total_amount_due = subtotal
    
    # Ensure total_amount_due is not negative
    total_amount_due = max(0, total_amount_due)

    story.append(Paragraph("<b>Sub-Totals</b>", styles['Heading3']))

    totals_data = [
        ['Mileage', f"${int(total_mileage)}"],
        ['Airfare', f"${total_airfare:.2f}"],
        ['Ground Transportation', f"${total_ground_transport:.2f}"],
        ['Parking', f"${total_parking:.2f}"],
        ['Lodging', f"${total_lodging:.2f}"],
        ['Baggage Fees', f"${total_baggage:.2f}"],
        ['Miscellaneous/Other', f"${total_misc:.2f}"],
        ['Per Diem', f"${total_per_diem:.2f}"],
        [Paragraph('<b>Total Amount Due</b>', styles['Normal']), Paragraph(f'<b>${total_amount_due:.2f}</b>', styles['Normal'])]
    ]
    
    totals_table = Table(totals_data, colWidths=[3*inch, 1.5*inch])
    totals_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.white),
        ('TEXTCOLOR', (1, 0), (1, 7), colors.red),
        ('BACKGROUND', (1, 0), (1, 7), colors.HexColor('#FFF5F5')),
        ('BACKGROUND', (0, 8), (0, 8), colors.white),
        ('BACKGROUND', (1, 8), (1, 8), colors.HexColor('#FFF5F5')),
        ('TEXTCOLOR', (1, 8), (1, 8), colors.red),
        ('FONTNAME', (0, 8), (0, 8), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.white]),
        ('TEXTCOLOR', (0, 9), (1, 9), colors.red),
        ('BACKGROUND', (0, 9), (1, 9), colors.white),
        ('FONTNAME', (0, 9), (1, 9), 'Helvetica-Bold'),
    ]))
    story.append(totals_table)
    story.append(Spacer(1, 0.15*inch))


    story.append(Paragraph("<b>Approval Signatures</b>", styles['Heading2']))
    # Signature section
    signature_text = form_data.get('signature', '').strip()
    
    # Create signature cell with image or text
    signature_cell_value = ''
    
    # Generate signature image from text
    if signature_text:
        try:
            # Generate signature image from text with high resolution (3x scale)
            signature_img_pil = generate_signature_image(signature_text, width=800, height=150, scale_factor=3)
            
            if signature_img_pil:
                # Ensure it's RGB with white background (blank)
                if signature_img_pil.mode != 'RGB':
                    rgb_img = PILImage.new('RGB', signature_img_pil.size, (255, 255, 255))
                    if signature_img_pil.mode == 'RGBA':
                        rgb_img.paste(signature_img_pil, mask=signature_img_pil.split()[3])
                    else:
                        rgb_img.paste(signature_img_pil)
                    signature_img_pil = rgb_img
                
                # Resize signature to fit the table cell (cell width is 2 inches, accounting for padding)
                # Cell has 6pt left/right padding, so available width is ~1.88 inches
                max_width = 1.88 * inch
                max_height = 0.5 * inch  # Reduced height to fit better in cell
                
                img_width, img_height = signature_img_pil.size
                aspect_ratio = img_height / img_width if img_width > 0 else 1
                
                # Calculate size maintaining aspect ratio but respecting both max width and height
                new_width = min(img_width, max_width)
                new_height = new_width * aspect_ratio
                
                if new_height > max_height:
                    new_height = max_height
                    new_width = new_height / aspect_ratio
                
                # Ensure width doesn't exceed cell width
                new_width = min(new_width, max_width)
                
                # Save to buffer for ReportLab with high quality
                img_buffer = io.BytesIO()
                # Save at full resolution for maximum clarity
                signature_img_pil.save(img_buffer, format='PNG', optimize=False, compress_level=1)
                img_buffer.seek(0)
                
                # Create ReportLab Image - use the calculated dimensions
                # The image will be high-res internally but displayed at the correct size
                signature_img = Image(img_buffer, width=new_width, height=new_height)
                signature_cell_value = signature_img
            else:
                signature_cell_value = signature_text
        except Exception as e:
            # Fallback to text if image generation fails
            signature_cell_value = signature_text
    
    # Combined Approval Signatures and Operations Use Only table
    # Use Paragraph for all labels to ensure consistent font size and width
    label_style = ParagraphStyle(
        'LabelStyle',
        parent=styles['Normal'],
        fontSize=9,
        fontName='Helvetica',
        alignment=0,  # LEFT
    )
    
    traveler_label = Paragraph("Traveler Signature", label_style)
    program_assistant_label = Paragraph("Program Assistant", label_style)
    lead_provider_text = Paragraph("Lead Technical\nAssistance Provider", label_style)
    
    combined_data = [
        [traveler_label, signature_cell_value, 'DATE', form_data.get('signature_date', '')],
        [program_assistant_label, '', 'DATE', ''],
        [lead_provider_text, '', 'DATE', ''],
        ['AWD', 'AWD-7776588', 'GR', 'GR426936'],
    ]
    
    combined_table = Table(combined_data, colWidths=[1.5*inch, 2*inch, 0.8*inch, 1.5*inch])
    combined_table.setStyle(TableStyle([
        # Grid and alignment for all rows
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        # Padding for all rows
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        # Traveler Signature row (row 0) - all white background, signature/date cells red text
        ('BACKGROUND', (0, 0), (-1, 0), colors.white),
        ('TEXTCOLOR', (1, 0), (1, 0), colors.red),
        ('TEXTCOLOR', (3, 0), (3, 0), colors.red),
        # Operations rows (rows 1-3) - all white background
        ('BACKGROUND', (0, 1), (-1, 3), colors.white),
    ]))
    story.append(combined_table)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer

def main():
    # Header with logo and title
    col_logo, col_title, col_spacer = st.columns([1,4,1], gap="large")
    with col_logo:
        try:
            georgetown_logo_url = 'https://raw.githubusercontent.com/JiaqinWu/HRSA64_Dash/main/Georgetown_logo_blueRGB.png'
            st.image(georgetown_logo_url, width=150)
        except:
            pass
    with col_title:
        st.markdown("<h1 style='text-align: center;'>Georgetown Domestic Travel Authorization Form Generator</h1>", unsafe_allow_html=True)
    with col_spacer:
        pass
    
    st.markdown("Fill out the form below to generate your Georgetown domestic travel authorization form.")
    
    try:
        wb, ws = load_excel_template()
        
        # General Guidance (UI only)
        with st.expander("General Guidance", expanded=False):
            st.markdown("""
            ### General Information
            Fill out the fields highlighted in green, as applicable. Form must be submitted at least one month prior to your proposed dates of travel. Please inform ADVANCE leadership if extenuating circumstances will prevent you from meeting this deadline.

            ### Receipts
            You must submit receipts as part of your Expense Report in GMS for every item associated with your trip. This signed travel authorization will serve as your receipt for meals and incidentals. Meals are reimbursed at the Federal Per Diem rate for the destination city.

            ### Mileage
            In lieu of taxi expenses, you can choose to be reimbursed for the mileage driven from your point of origin to the airport, train station, or bus station. Georgetown University uses the IRS mileage rate.

            Please attach documentation for the specified mileage in your GMS Expense report (e.g., Google Maps, MapQuest). Round all mileage to the nearest mile.

            ### Airfare, Transportation, Parking, Lodging, Baggage Fees, Miscellaneous/Other
            - **Airfare**: Should be booked through Concur and paid by Georgetown University. Include it as a cost in this Travel Authorization Form; your airfare should be included as an expense in your GMS Expense Report, but not as a personal reimbursement. If you are being reimbursed for your air travel, you must submit your itinerary and receipt.
            - **Ground Transportation**: Covers reasonable expenses for taxis or other modes of transportation to and from airports and/or train and bus stations. Receipts must indicate the point of departure and point of arrival.
            - **Parking**: If you are being reimbursed for parking, you must submit your receipt(s).
            - **Lodging**: If lodging is purchased by the traveler, hotel receipts must be submitted. Lodging includes room and tax; it does not include telephone calls, room service, or other incidentals.
            - **Baggage Fees**: Georgetown University will reimburse for one checked bag per passenger for each leg of trip (if the carrier charges for checked bags). For carriers with a free first bag, no reimbursement for additional bags will be allowed.
            - **Miscellaneous/Other**: Includes pre‑approved travel expenses not listed in this form.

            ### Meals and Incidental Expenses (M&IE)
            Georgetown University will reimburse meals and incidentals at the U.S. Government per diem rates. This allowance covers tips, porter fees, etc.

            Federal Guidelines stipulate that on the first and last travel day, travelers are only eligible for 75 percent of the total M&IE rate.

            The cost of any meals provided at meetings and conferences will not be reimbursed by Georgetown University. For meals that have been provided by Georgetown University, please place an "x" in the appropriate box on the reimbursement form.
            """)
        
        # Date inputs outside form so they trigger immediate reruns
        st.header("Travel Dates")
        col_date1, col_date2 = st.columns(2)
        with col_date1:
            departure_date = st.date_input("Departure Date *", key="departure_date")
        with col_date2:
            return_date = st.date_input("Return Date *", key="return_date")
        
        # Validate date range
        if departure_date and return_date and return_date < departure_date:
            st.error("Return Date must be the same as or after the Departure Date.")
            st.stop()
        
        # Track date changes to auto-populate date fields
        # Initialize session state for date tracking
        if 'last_departure' not in st.session_state:
            st.session_state.last_departure = departure_date
        if 'last_return' not in st.session_state:
            st.session_state.last_return = return_date
        
        # Check if dates changed
        dates_changed = (departure_date != st.session_state.last_departure or 
                       return_date != st.session_state.last_return)
        
        # Compute total days and generate full date range
        if departure_date and return_date and return_date >= departure_date:
            total_days = (return_date - departure_date).days + 1
            # Reasonable upper bound to avoid runaway UI
            total_days = min(total_days, 60)
        else:
            total_days = 7
        default_dates = generate_date_range(departure_date, return_date, max_days=total_days)
        
        # Update session state when dates change (this happens on rerun)
        if dates_changed:
            st.session_state.last_departure = departure_date
            st.session_state.last_return = return_date
            # Update all date fields with new defaults when dates change
            # Clear previous keys generously then set new defaults
            for i in range(0, 100):
                if i < len(default_dates) and default_dates[i]:
                    st.session_state[f'mileage_date_{i}'] = default_dates[i]
                    st.session_state[f'expense_date_{i}'] = default_dates[i]
                    st.session_state[f'per_diem_date_{i}'] = default_dates[i]
                else:
                    # Clear if beyond date range
                    st.session_state[f'mileage_date_{i}'] = ''
                    st.session_state[f'expense_date_{i}'] = ''
                    st.session_state[f'per_diem_date_{i}'] = ''
        else:
            # Initialize session state on first load if not exists
            for i in range(total_days):
                if f'mileage_date_{i}' not in st.session_state:
                    st.session_state[f'mileage_date_{i}'] = default_dates[i] if i < len(default_dates) else ''
                if f'expense_date_{i}' not in st.session_state:
                    st.session_state[f'expense_date_{i}'] = default_dates[i] if i < len(default_dates) else ''
                if f'per_diem_date_{i}' not in st.session_state:
                    st.session_state[f'per_diem_date_{i}'] = default_dates[i] if i < len(default_dates) else ''
        
        with st.form("travel_form"):
            st.header("Traveler Information")
            col1, col2 = st.columns(2)
            
            with col1:
                name = st.text_input("Name *", key="name")
                organization = st.text_input("Organization", value="Georgetown University", key="organization")
                destination = st.text_input("Destination *", key="destination")
                email = st.text_input("Email Address *", key="email")              
                
            
            with col2:
                address1 = st.text_input("Address Line 1 *", key="address1")
                address2 = st.text_input("Address Line 2", key="address2")
                city = st.text_input("City *", key="city")
                state = st.text_input("State *", key="state")
                zip_code = st.text_input("Zip *", key="zip")

            st.header("Purpose of Travel")
            
            st.header("Mileage Expenses")
            st.markdown("**Mileage rate for 2025: $0.70 per mile**")
            
            mileage_dates = []
            mileage_amounts = []
            
            # Render mileage inputs in chunks of 7 days per row
            for chunk_start in range(0, total_days, 7):
                chunk_len = min(7, total_days - chunk_start)
                cols = st.columns(chunk_len)
                for offset in range(chunk_len):
                    i = chunk_start + offset
                    with cols[offset]:
                        mileage_dates.append(st.text_input(f"Day {i+1}", key=f"mileage_date_{i}", placeholder="MM/DD/YY"))
                        mileage_amounts.append(number_text_input(f"Miles", key=f"mileage_{i}", value=0.0, placeholder="0"))
            
            total_mileage = round(sum([m * 0.70 for m in mileage_amounts if m]),0)
            
            st.header("Travel Expenses")
            expense_dates = []
            airfare = []
            ground_transport = []
            parking = []
            lodging = []
            baggage = []
            misc = []
            misc2 = []
            # First pass: render Date, Airfare, Ground, Parking, Lodging, Baggage
            for chunk_start in range(0, total_days, 7):
                chunk_len = min(7, total_days - chunk_start)
                cols = st.columns(chunk_len)
                for offset in range(chunk_len):
                    i = chunk_start + offset
                    with cols[offset]:
                        expense_dates.append(st.text_input(f"Day {i+1}", key=f"expense_date_{i}", placeholder="MM/DD/YY"))
                        airfare.append(number_text_input(f"Airfare", key=f"airfare_{i}", value=0.0, placeholder="0.00"))
                        ground_transport.append(number_text_input(f"Ground Transportation", key=f"ground_{i}", value=0.0, placeholder="0.00"))
                        parking.append(number_text_input(f"Parking", key=f"parking_{i}", value=0.0, placeholder="0.00"))
                        lodging.append(number_text_input(f"Lodging", key=f"lodging_{i}", value=0.0, placeholder="0.00"))
                        baggage.append(number_text_input(f"Baggage Fees", key=f"baggage_{i}", value=0.0, placeholder="0.00"))

            # Descriptions next (always shown above misc rows, once for the section)
            misc_desc1 = st.text_input("Miscellaneous/Other Description 1", key="misc_desc1", placeholder="e.g., Registration")
            # Second pass: render Misc Row 1 and Misc Row 2 amounts
            for chunk_start in range(0, total_days, 7):
                chunk_len = min(7, total_days - chunk_start)
                cols = st.columns(chunk_len)
                for offset in range(chunk_len):
                    i = chunk_start + offset
                    with cols[offset]:
                        misc.append(number_text_input(f"{misc_desc1} Day {i+1}", key=f"misc_{i}", value=0.0, placeholder="0.00"))

            misc_desc2 = st.text_input("Miscellaneous/Other Description 2", key="misc_desc2", placeholder="e.g., Supplies")

            # Second pass: render Misc Row 1 and Misc Row 2 amounts
            for chunk_start in range(0, total_days, 7):
                chunk_len = min(7, total_days - chunk_start)
                cols = st.columns(chunk_len)
                for offset in range(chunk_len):
                    i = chunk_start + offset
                    with cols[offset]:
                        misc2.append(number_text_input(f"{misc_desc2} Day {i+1}", key=f"misc2_{i}", value=0.0, placeholder="0.00"))
            
            
            st.header("Meals and Incidentals Per Diem")
            st.markdown("**Please confirm the official GSA per diem rate for your travel destination at https://www.gsa.gov/travel/plan-book/per-diem-rates and select the corresponding rate below.**")
            # Single per diem selection for all days
            selected_per_diem = st.selectbox("Per Diem Rate (applies to all days)", options=[68,74,80,86,92], index=2, key="per_diem_base")
            per_diem_dates = []
            per_diem_amounts = []
            breakfast_checks = []
            lunch_checks = []
            dinner_checks = []
            st.markdown("**Check boxes if meals were provided by Georgetown University**")
            # Render per diem inputs in chunks of 7 days per row
            for chunk_start in range(0, total_days, 7):
                chunk_len = min(7, total_days - chunk_start)
                cols = st.columns(chunk_len)
                for offset in range(chunk_len):
                    i = chunk_start + offset
                    with cols[offset]:
                        per_diem_dates.append(st.text_input(f"Day {i+1}", key=f"per_diem_date_{i}", placeholder="MM/DD/YY"))
                        per_diem_amounts.append(selected_per_diem) 
                        breakfast_checks.append(st.checkbox(f"Breakfast", key=f"breakfast_{i}"))
                        lunch_checks.append(st.checkbox(f"Lunch", key=f"lunch_{i}"))
                        dinner_checks.append(st.checkbox(f"Dinner", key=f"dinner_{i}"))
            
            st.header("Additional Information")
            
            # E-Signature section
            st.subheader("Traveler Signature")
            col1, col2 = st.columns([2, 1])
            with col1:
                signature_text = st.text_input("Type your full name", key="signature_text", 
                                              help="Your typed name will be automatically converted to a signature-style image")
                if signature_text:
                    # Show preview of signature (use lower scale for preview to be faster)
                    try:
                        preview_img = generate_signature_image(signature_text, width=600, height=120, scale_factor=2)
                        if preview_img:
                            # Ensure it's RGB for display (should already be RGB now)
                            if preview_img.mode != 'RGB':
                                rgb_preview = PILImage.new('RGB', preview_img.size, (255, 255, 255))
                                if preview_img.mode == 'RGBA':
                                    rgb_preview.paste(preview_img, mask=preview_img.split()[3])
                                else:
                                    rgb_preview.paste(preview_img)
                                preview_img = rgb_preview
                            # Resize preview for display
                            preview_display = preview_img.resize((400, int(400 * preview_img.size[1] / preview_img.size[0])))
                            st.image(preview_display, caption="Signature Preview", width=400)
                    except Exception as e:
                        pass
            with col2:
                signature_date = st.date_input("Signature Date", value=datetime.now().date(), key="sig_date")
            
            signature = signature_text.strip() if signature_text else ""
            
            submitted = st.form_submit_button("Generate PDF")
        
        if submitted:
            # Validate required Traveler Information fields
            missing_fields = []
            if not name or not name.strip():
                missing_fields.append("Name")
            if not address1 or not address1.strip():
                missing_fields.append("Address Line 1")
            if not city or not city.strip():
                missing_fields.append("City")
            if not state or not state.strip():
                missing_fields.append("State")
            if not zip_code or not zip_code.strip():
                missing_fields.append("Zip")
            if not destination or not destination.strip():
                missing_fields.append("Destination")
            if not email or not email.strip():
                missing_fields.append("Email Address")
            
            if missing_fields:
                st.warning(f"⚠️ Please fill in all required fields: {', '.join(missing_fields)}")
                st.stop()
            
            # Check for any input validation errors (check all number inputs)
            has_validation_errors = False
            # Check all input keys that might have errors
            input_prefixes = ['mileage_', 'airfare_', 'ground_', 'parking_', 'lodging_', 'baggage_', 'misc_', 'misc2_']
            for key in st.session_state.keys():
                if key.endswith('_has_error') and st.session_state[key]:
                    # Check if this is one of our input fields
                    base_key = key.replace('_has_error', '')
                    if any(base_key.startswith(prefix) for prefix in input_prefixes):
                        has_validation_errors = True
                        break
            
            if has_validation_errors:
                st.warning("⚠️ **Cannot generate PDF: Please fix all invalid input fields above.**")
                st.stop()
            # Calculate totals
            total_airfare = sum(airfare)
            total_ground_transport = sum(ground_transport)
            total_parking = sum(parking)
            total_lodging = sum(lodging)
            total_baggage = sum(baggage)
            total_misc = sum(misc) + sum(misc2)  # Include both misc rows in total
            # Calculate adjusted per diem with meal deductions
            days_with_dates = [i for i, d in enumerate(per_diem_dates) if d and str(d).strip()]
            num_days = len(days_with_dates)
            first_day_idx = days_with_dates[0] if days_with_dates else 0
            last_day_idx = days_with_dates[-1] if days_with_dates else 0
            
            meal_deductions = {
                68: { 'breakfast': 16, 'lunch': 19, 'dinner': 28, 'incidental': 5, 'first_last': 51.00 },
                74: { 'breakfast': 18, 'lunch': 20, 'dinner': 31, 'incidental': 5, 'first_last': 55.50 },
                80: { 'breakfast': 20, 'lunch': 22, 'dinner': 33, 'incidental': 5, 'first_last': 60.00 },
                86: { 'breakfast': 22, 'lunch': 23, 'dinner': 36, 'incidental': 5, 'first_last': 64.50 },
                92: { 'breakfast': 23, 'lunch': 26, 'dinner': 38, 'incidental': 5, 'first_last': 69.00 },
            }
            adjusted_per_diem_daily = []
            for i in range(len(per_diem_dates)):
                if i < len(per_diem_dates) and per_diem_dates[i] and str(per_diem_dates[i]).strip():
                    base_per_diem = int(per_diem_amounts[i]) if (i < len(per_diem_amounts) and per_diem_amounts[i]) else 80
                    deducts = meal_deductions.get(base_per_diem, meal_deductions[80])
                    deduction_total = 0.0
                    if i < len(breakfast_checks) and breakfast_checks[i]:
                        deduction_total += deducts['breakfast']
                    if i < len(lunch_checks) and lunch_checks[i]:
                        deduction_total += deducts['lunch']
                    if i < len(dinner_checks) and dinner_checks[i]:
                        deduction_total += deducts['dinner']
                    # Base already includes incidentals; do not add +$5 here
                    pre75_total = max(0.0, float(base_per_diem) - deduction_total)
                    # Apply 75% for first and last day
                    if i == first_day_idx or i == last_day_idx:
                        final_per_diem = round(pre75_total * 0.75, 2)
                    else:
                        final_per_diem = round(pre75_total, 2)
                    
                    adjusted_per_diem_daily.append(final_per_diem)
                else:
                    adjusted_per_diem_daily.append(0.0)
            
            total_per_diem = sum(adjusted_per_diem_daily)
            total_amount_due = (total_mileage + total_airfare + total_ground_transport + 
                              total_parking + total_lodging + total_baggage + 
                              total_misc + total_per_diem)
            
            form_data = {
                'name': name,
                'address1': address1,
                'address2': address2,
                'city': city,
                'state': state,
                'zip': zip_code,
                'organization': organization,
                'destination': destination,
                'departure_date': departure_date.strftime('%m/%d/%Y') if departure_date else '',
                'return_date': return_date.strftime('%m/%d/%Y') if return_date else '',
                'email': email,
                'mileage_dates': mileage_dates,
                'mileage_amounts': mileage_amounts,
                'total_mileage': total_mileage,
                'expense_dates': expense_dates,
                'airfare': airfare,
                'ground_transport': ground_transport,
                'parking': parking,
                'lodging': lodging,
                'baggage': baggage,
                'misc': misc,
                'misc2': misc2,
                'misc_desc1': misc_desc1,
                'misc_desc2': misc_desc2,
                'total_airfare': total_airfare,
                'total_ground_transport': total_ground_transport,
                'total_parking': total_parking,
                'total_lodging': total_lodging,
                'total_baggage': total_baggage,
                'total_misc': total_misc,
                'per_diem_dates': per_diem_dates,
                'per_diem_amounts': per_diem_amounts,
                'breakfast_checks': breakfast_checks,
                'lunch_checks': lunch_checks,
                'dinner_checks': dinner_checks,
                'total_per_diem': total_per_diem,
                'total_amount_due': total_amount_due,
                'signature': signature,
                'signature_date': signature_date.strftime('%m/%d/%Y') if signature_date else ''
            }
            # Store for review step
            st.session_state['review_data'] = form_data
            st.success("Please review the information below and approve to finalize.")

        # Review & Approve pane
        if 'review_data' in st.session_state:
            review = st.session_state['review_data']
            st.subheader("Review & Approve")
            colA, colB = st.columns(2)
            with colA:
                st.markdown("**Traveler**")
                traveler_html = f"""
                <div style='border:1px solid #e0e0e0;border-radius:8px;padding:12px;background:#fafafa;'>
                  <div style='display:flex;justify-content:space-between;padding:4px 0;'>
                    <span style='color:#555;'>Name</span><strong>{review.get('name','')}</strong>
                  </div>
                  <div style='display:flex;justify-content:space-between;padding:4px 0;'>
                    <span style='color:#555;'>Organization</span><strong>{review.get('organization','')}</strong>
                  </div>
                  <div style='display:flex;justify-content:space-between;padding:4px 0;'>
                    <span style='color:#555;'>Destination</span><strong>{review.get('destination','')}</strong>
                  </div>
                  <div style='display:flex;justify-content:space-between;padding:4px 0;'>
                    <span style='color:#555;'>Email</span><strong>{review.get('email','')}</strong>
                  </div>
                </div>
                """
                st.markdown(traveler_html, unsafe_allow_html=True)
            with colB:
                st.markdown("**Trip**")
                trip_html = f"""
                <div style='border:1px solid #e0e0e0;border-radius:8px;padding:12px;background:#fafafa;'>
                  <div style='display:flex;justify-content:space-between;padding:4px 0;'>
                    <span style='color:#555;'>Departure Date</span><strong>{review.get('departure_date','')}</strong>
                  </div>
                  <div style='display:flex;justify-content:space-between;padding:4px 0;'>
                    <span style='color:#555;'>Return Date</span><strong>{review.get('return_date','')}</strong>
                  </div>
                </div>
                """
                st.markdown(trip_html, unsafe_allow_html=True)
            st.markdown("**Totals**")
            totals_html = f"""
            <table style='width:100%;border-collapse:collapse;border:1px solid #eee;'>
              <thead>
                <tr style='background:#f5f5f5;'>
                  <th style='text-align:left;padding:8px;border-bottom:1px solid #eee;'>Category</th>
                  <th style='text-align:right;padding:8px;border-bottom:1px solid #eee;'>Amount</th>
                </tr>
              </thead>
              <tbody>
                <tr><td style='padding:8px;border-bottom:1px solid #f0f0f0;'>Mileage</td><td style='padding:8px;text-align:right;'>${int(review.get('total_mileage',0))}</td></tr>
                <tr><td style='padding:8px;border-bottom:1px solid #f0f0f0;'>Airfare</td><td style='padding:8px;text-align:right;'>${review.get('total_airfare',0):.2f}</td></tr>
                <tr><td style='padding:8px;border-bottom:1px solid #f0f0f0;'>Ground Transport</td><td style='padding:8px;text-align:right;'>${review.get('total_ground_transport',0):.2f}</td></tr>
                <tr><td style='padding:8px;border-bottom:1px solid #f0f0f0;'>Parking</td><td style='padding:8px;text-align:right;'>${review.get('total_parking',0):.2f}</td></tr>
                <tr><td style='padding:8px;border-bottom:1px solid #f0f0f0;'>Lodging</td><td style='padding:8px;text-align:right;'>${review.get('total_lodging',0):.2f}</td></tr>
                <tr><td style='padding:8px;border-bottom:1px solid #f0f0f0;'>Baggage</td><td style='padding:8px;text-align:right;'>${review.get('total_baggage',0):.2f}</td></tr>
                <tr><td style='padding:8px;border-bottom:1px solid #f0f0f0;'>Miscellaneous</td><td style='padding:8px;text-align:right;'>${review.get('total_misc',0):.2f}</td></tr>
                <tr><td style='padding:8px;border-bottom:1px solid #f0f0f0;'>Per Diem</td><td style='padding:8px;text-align:right;'>${review.get('total_per_diem',0):.2f}</td></tr>
                <tr style='background:#fff8f8;font-weight:600;'>
                  <td style='padding:8px;border-top:1px solid #eee;'>Total Amount Due</td>
                  <td style='padding:8px;text-align:right;border-top:1px solid #eee;'>${review.get('total_amount_due',0):.2f}</td>
                </tr>
              </tbody>
            </table>
            """
            st.markdown(totals_html, unsafe_allow_html=True)
            approved = st.checkbox("I have reviewed and approve this travel form.", key="approve_review")
            generate_now = st.button("Finalize and Download PDF", disabled=not approved)
            if generate_now and approved:
                pdf_buffer = create_pdf(review, ws)
                st.success("✅ PDF generated successfully!")
                st.download_button(
                    label="📥 Download PDF",
                    data=pdf_buffer,
                    file_name=f"Travel_Authorization_Form_{review.get('name','')}_{review.get('departure_date','')}_{review.get('return_date','')}.pdf",
                    mime="application/pdf"
                )
                # Clear review data after generation
                del st.session_state['review_data']
    
    except Exception as e:
        st.error(f"Error: {str(e)}")
        st.exception(e)

if __name__ == "__main__":
    main()
